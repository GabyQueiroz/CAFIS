from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import io
import os
import secrets
import smtplib
import sqlite3
from datetime import date, datetime, timedelta, timezone
from email.message import EmailMessage
from pathlib import Path
import re
import unicodedata
from typing import Any

import qrcode
from openpyxl import load_workbook
from fastapi import Cookie, Depends, FastAPI, File, Form, HTTPException, Response, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.getenv("CAFIS_DB_PATH", BASE_DIR / "cafis_academia.db"))
FRONTEND_DIR = BASE_DIR / "frontend"
UTC = timezone.utc
COOKIE_NAME = "cafis_session"
SESSION_SECONDS = 60 * 60 * 10
PBKDF2_ITERS = 260_000
COOKIE_SECURE = os.getenv("CAFIS_COOKIE_SECURE", "0").lower() in {"1", "true", "yes"}
WEEKDAY_INDEX = {"segunda": 0, "terca": 1, "terça": 1, "quarta": 2, "quinta": 3, "sexta": 4, "sabado": 5, "sábado": 5}
IMPORT_SHEETS = ("Segunda", "Terça", "Quarta", "Quinta", "Sexta")

app = FastAPI(title="CAFIS Academia UTFPR PG", version="1.0.0")


def now_iso() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat()


def connect() -> sqlite3.Connection:
    db_path = DB_PATH
    try:
        db_path.parent.mkdir(parents=True, exist_ok=True)
    except PermissionError:
        db_path = Path(os.getenv("CAFIS_FALLBACK_DB_PATH", "/tmp/cafis_academia.db"))
        db_path.parent.mkdir(parents=True, exist_ok=True)
        print(f"WARNING: sem permissao para {DB_PATH}; usando banco temporario em {db_path}")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def hash_password(password: str, salt: str | None = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), PBKDF2_ITERS)
    return f"pbkdf2_sha256${PBKDF2_ITERS}${salt}${digest.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        algo, iters, salt, expected = stored.split("$")
        if algo != "pbkdf2_sha256":
            return False
        digest = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), int(iters)).hex()
        return hmac.compare_digest(digest, expected)
    except Exception:
        return False


def normalize_email(email: str) -> str:
    return email.strip().lower()


def rowdict(row: sqlite3.Row | None) -> dict[str, Any] | None:
    return dict(row) if row else None


def normalize_text(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch)).lower()


def clean_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def only_digits(value: Any) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def as_iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_label(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def parse_time_range(text: str) -> tuple[str, str]:
    matches = re.findall(r"(\d{1,2}[:h]\d{2})", text.lower())
    if len(matches) >= 2:
        start, end = matches[0], matches[1]
        return start.replace("h", ":"), end.replace("h", ":")
    compact = re.findall(r"(\d{1,2}:\d{2})", text)
    if len(compact) >= 2:
        return compact[0], compact[1]
    return "", ""


def canonical_weekday_label(value: str) -> str:
    normalized = normalize_text(value)
    return {
        "segunda": "Segunda",
        "terca": "Terça",
        "quarta": "Quarta",
        "quinta": "Quinta",
        "sexta": "Sexta",
        "sabado": "Sábado",
    }.get(normalized, clean_label(value))


def daterange(start_date: date, end_date: date):
    cursor = start_date
    while cursor <= end_date:
        yield cursor
        cursor += timedelta(days=1)


def ensure_admin_user(
    db: sqlite3.Connection,
    name: str,
    email: str,
    password: str,
    notes: str,
    update_password: bool = False,
) -> None:
    admin_email = normalize_email(email)
    existing = db.execute("SELECT id FROM users WHERE email = ?", (admin_email,)).fetchone()
    if existing:
        if update_password:
            db.execute(
                "UPDATE users SET password_hash=?, active=1, role='admin' WHERE id=?",
                (hash_password(password), existing["id"]),
            )
        return
    db.execute(
        """
        INSERT INTO users
        (name, email, password_hash, role, qr_secret, created_at, notes)
        VALUES (?, ?, ?, 'admin', ?, ?, ?)
        """,
        (
            name,
            admin_email,
            hash_password(password),
            secrets.token_urlsafe(32),
            now_iso(),
            notes,
        ),
    )


def init_db() -> None:
    with connect() as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin','student')),
                cpf TEXT,
                registration TEXT,
                birth_date TEXT,
                phone TEXT,
                goal TEXT DEFAULT 'saude',
                availability_days TEXT DEFAULT 'seg,qua,sex',
                weekly_minutes INTEGER DEFAULT 180,
                notes TEXT DEFAULT '',
                qr_secret TEXT NOT NULL UNIQUE,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS sessions (
                token_hash TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                expires_at INTEGER NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS evaluations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                evaluation_type TEXT NOT NULL DEFAULT 'completa',
                created_at TEXT NOT NULL,
                weight REAL,
                height_cm REAL,
                body_fat REAL,
                body_water REAL,
                muscle_mass REAL,
                bmr REAL,
                metabolic_age REAL,
                bone_mass REAL,
                visceral_fat REAL,
                waist_cm REAL,
                hip_cm REAL,
                flexibility_cm REAL,
                abdominal_reps INTEGER,
                pushup_reps INTEGER,
                resting_hr INTEGER,
                post_hr INTEGER,
                recovery_hr_5min INTEGER,
                cooper_km REAL,
                vo2max REAL,
                systolic INTEGER,
                diastolic INTEGER,
                qualia_score REAL NOT NULL,
                risk_level TEXT NOT NULL,
                recommendations TEXT NOT NULL,
                notes TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                check_in TEXT NOT NULL,
                check_out TEXT,
                minutes INTEGER DEFAULT 0,
                registered_by INTEGER REFERENCES users(id),
                source TEXT DEFAULT 'qr'
            );
            CREATE TABLE IF NOT EXISTS equipment (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'ok',
                quantity INTEGER NOT NULL DEFAULT 1,
                location TEXT DEFAULT '',
                photo_data_url TEXT DEFAULT '',
                maintenance_notes TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS workout_plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                month TEXT NOT NULL,
                week INTEGER NOT NULL,
                plan_text TEXT NOT NULL,
                created_by INTEGER REFERENCES users(id),
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sender_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                recipient_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                subject TEXT NOT NULL,
                body TEXT NOT NULL,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS programs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                slug TEXT NOT NULL UNIQUE,
                description TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS class_groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                program_id INTEGER NOT NULL REFERENCES programs(id) ON DELETE CASCADE,
                name TEXT NOT NULL,
                schedule TEXT DEFAULT '',
                teacher TEXT DEFAULT '',
                default_minutes INTEGER NOT NULL DEFAULT 60,
                location TEXT DEFAULT '',
                weekday TEXT DEFAULT '',
                start_time TEXT DEFAULT '',
                end_time TEXT DEFAULT '',
                activity_start TEXT DEFAULT '',
                activity_end TEXT DEFAULT '',
                academic_year INTEGER DEFAULT 0,
                period_label TEXT DEFAULT '',
                source_file TEXT DEFAULT '',
                event_code TEXT DEFAULT '',
                notes TEXT DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS enrollments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                class_id INTEGER NOT NULL REFERENCES class_groups(id) ON DELETE CASCADE,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                created_at TEXT NOT NULL,
                UNIQUE(class_id, user_id)
            );
            CREATE TABLE IF NOT EXISTS class_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                class_id INTEGER NOT NULL REFERENCES class_groups(id) ON DELETE CASCADE,
                session_date TEXT NOT NULL,
                notes TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                UNIQUE(class_id, session_date)
            );
            CREATE TABLE IF NOT EXISTS class_attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id INTEGER NOT NULL REFERENCES class_sessions(id) ON DELETE CASCADE,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                status TEXT NOT NULL CHECK(status IN ('present','absent','justified')),
                minutes INTEGER NOT NULL DEFAULT 0,
                notes TEXT DEFAULT '',
                registered_at TEXT NOT NULL,
                UNIQUE(session_id, user_id)
            );
            """
        )
        for name, slug, description in [
            ("Academia", "academia", "Controle de alunos, avaliações, presenças e horas da academia CAFIS."),
            ("Natação", "natacao", "Controle de turmas, chamadas, horas e certificados do projeto de natação."),
        ]:
            if not db.execute("SELECT id FROM programs WHERE slug=?", (slug,)).fetchone():
                db.execute(
                    "INSERT INTO programs (name, slug, description, created_at) VALUES (?, ?, ?, ?)",
                    (name, slug, description, now_iso()),
                )
        eval_cols = db.execute("PRAGMA table_info(evaluations)").fetchall()
        eval_info = {row["name"]: row for row in eval_cols}
        weight_col = eval_info.get("weight")
        height_col = eval_info.get("height_cm")
        needs_eval_rebuild = (
            "evaluation_type" not in eval_info
            or bool(weight_col["notnull"] if weight_col else False)
            or bool(height_col["notnull"] if height_col else False)
        )
        if needs_eval_rebuild:
            db.executescript(
                """
                CREATE TABLE IF NOT EXISTS evaluations_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    evaluation_type TEXT NOT NULL DEFAULT 'completa',
                    created_at TEXT NOT NULL,
                    weight REAL,
                    height_cm REAL,
                    body_fat REAL,
                    body_water REAL,
                    muscle_mass REAL,
                    bmr REAL,
                    metabolic_age REAL,
                    bone_mass REAL,
                    visceral_fat REAL,
                    waist_cm REAL,
                    hip_cm REAL,
                    flexibility_cm REAL,
                    abdominal_reps INTEGER,
                    pushup_reps INTEGER,
                    resting_hr INTEGER,
                    post_hr INTEGER,
                    recovery_hr_5min INTEGER,
                    cooper_km REAL,
                    vo2max REAL,
                    systolic INTEGER,
                    diastolic INTEGER,
                    qualia_score REAL NOT NULL,
                    risk_level TEXT NOT NULL,
                    recommendations TEXT NOT NULL,
                    notes TEXT DEFAULT ''
                );
                INSERT INTO evaluations_new
                (id,user_id,evaluation_type,created_at,weight,height_cm,body_fat,body_water,muscle_mass,bmr,metabolic_age,bone_mass,visceral_fat,
                 waist_cm,hip_cm,flexibility_cm,abdominal_reps,pushup_reps,resting_hr,post_hr,recovery_hr_5min,cooper_km,
                 vo2max,systolic,diastolic,qualia_score,risk_level,recommendations,notes)
                SELECT id,user_id,'completa',created_at,weight,height_cm,body_fat,body_water,muscle_mass,bmr,metabolic_age,bone_mass,visceral_fat,
                 waist_cm,hip_cm,flexibility_cm,abdominal_reps,pushup_reps,resting_hr,post_hr,recovery_hr_5min,cooper_km,
                 vo2max,systolic,diastolic,qualia_score,risk_level,recommendations,notes
                FROM evaluations;
                DROP TABLE evaluations;
                ALTER TABLE evaluations_new RENAME TO evaluations;
                """
            )
        for sql in [
            "ALTER TABLE equipment ADD COLUMN quantity INTEGER NOT NULL DEFAULT 1",
            "ALTER TABLE equipment ADD COLUMN photo_data_url TEXT DEFAULT ''",
            "ALTER TABLE class_groups ADD COLUMN weekday TEXT DEFAULT ''",
            "ALTER TABLE class_groups ADD COLUMN start_time TEXT DEFAULT ''",
            "ALTER TABLE class_groups ADD COLUMN end_time TEXT DEFAULT ''",
            "ALTER TABLE class_groups ADD COLUMN activity_start TEXT DEFAULT ''",
            "ALTER TABLE class_groups ADD COLUMN activity_end TEXT DEFAULT ''",
            "ALTER TABLE class_groups ADD COLUMN academic_year INTEGER DEFAULT 0",
            "ALTER TABLE class_groups ADD COLUMN period_label TEXT DEFAULT ''",
            "ALTER TABLE class_groups ADD COLUMN source_file TEXT DEFAULT ''",
            "ALTER TABLE class_groups ADD COLUMN event_code TEXT DEFAULT ''",
        ]:
            try:
                db.execute(sql)
            except sqlite3.OperationalError:
                pass
        ensure_admin_user(
            db,
            "Admin CAFIS",
            os.getenv("CAFIS_ADMIN_EMAIL", "admin@cafis.utfpr.edu.br"),
            os.getenv("CAFIS_ADMIN_PASSWORD", "Admin@12345"),
            "Troque a senha antes de publicar.",
            update_password=bool(os.getenv("CAFIS_ADMIN_PASSWORD")),
        )
        ensure_admin_user(
            db,
            "Administrador CAFIS",
            os.getenv("CAFIS_CAFIS_ADMIN_EMAIL", "adm.cafis@utfpr.edu.br"),
            os.getenv("CAFIS_CAFIS_ADMIN_PASSWORD", "CAFIS@2026"),
            "Administrador de bootstrap do CAFIS. Troque a senha no Render.",
            update_password=bool(os.getenv("CAFIS_CAFIS_ADMIN_PASSWORD")),
        )
        for name, category in [
            ("Esteira", "cardio"),
            ("Bicicleta ergometrica", "cardio"),
            ("Leg press", "forca"),
            ("Puxador alto", "forca"),
            ("Supino", "forca"),
            ("Halteres", "peso livre"),
        ]:
            if not db.execute("SELECT id FROM equipment WHERE name = ?", (name,)).fetchone():
                db.execute(
                    "INSERT INTO equipment (name, category, status, created_at) VALUES (?, ?, 'ok', ?)",
                    (name, category, now_iso()),
                )


@app.on_event("startup")
def startup() -> None:
    init_db()


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


class LoginIn(BaseModel):
    email: str
    password: str = Field(min_length=4, max_length=128)


class StudentIn(BaseModel):
    name: str = Field(min_length=3, max_length=120)
    email: str
    cpf: str = ""
    registration: str = ""
    birth_date: str = ""
    phone: str = ""
    goal: str = "saude"
    availability_days: str = "seg,qua,sex"
    weekly_minutes: int = Field(default=180, ge=30, le=900)
    notes: str = ""
    password: str | None = Field(default=None, min_length=6, max_length=128)


class PublicRegisterIn(BaseModel):
    name: str = Field(min_length=3, max_length=120)
    email: str
    cpf: str = Field(min_length=11, max_length=20)
    registration: str = Field(min_length=3, max_length=40)
    birth_date: str = ""
    phone: str = ""
    goal: str = "saude"
    availability_days: str = "seg,qua,sex"
    weekly_minutes: int = Field(default=180, ge=30, le=900)
    password: str = Field(min_length=8, max_length=128)


class EvaluationIn(BaseModel):
    user_id: int | None = None
    evaluation_type: str = "bio"
    weight: float | None = Field(default=None, gt=20, le=300)
    height_cm: float | None = Field(default=None, gt=100, le=240)
    body_fat: float | None = Field(default=None, ge=0, le=80)
    body_water: float | None = Field(default=None, ge=0, le=100)
    muscle_mass: float | None = Field(default=None, ge=0, le=100)
    bmr: float | None = Field(default=None, ge=0, le=6000)
    metabolic_age: float | None = Field(default=None, ge=0, le=120)
    bone_mass: float | None = Field(default=None, ge=0, le=20)
    visceral_fat: float | None = Field(default=None, ge=0, le=40)
    waist_cm: float | None = Field(default=None, ge=30, le=220)
    hip_cm: float | None = Field(default=None, ge=30, le=220)
    flexibility_cm: float | None = Field(default=None, ge=0, le=100)
    abdominal_reps: int | None = Field(default=None, ge=0, le=300)
    pushup_reps: int | None = Field(default=None, ge=0, le=300)
    resting_hr: int | None = Field(default=None, ge=30, le=220)
    post_hr: int | None = Field(default=None, ge=30, le=240)
    recovery_hr_5min: int | None = Field(default=None, ge=30, le=240)
    cooper_km: float | None = Field(default=None, ge=0, le=10)
    vo2max: float | None = Field(default=None, ge=0, le=100)
    systolic: int | None = Field(default=None, ge=60, le=260)
    diastolic: int | None = Field(default=None, ge=30, le=180)
    notes: str = ""


class AttendanceScanIn(BaseModel):
    qr_payload: str


class AttendanceManualIn(BaseModel):
    user_id: int


class EquipmentIn(BaseModel):
    name: str = Field(min_length=2, max_length=120)
    category: str = Field(min_length=2, max_length=80)
    status: str = "ok"
    quantity: int = Field(default=1, ge=0, le=999)
    location: str = ""
    photo_data_url: str = Field(default="", max_length=2_000_000)
    maintenance_notes: str = ""


class MessageIn(BaseModel):
    subject: str = Field(min_length=3, max_length=180)
    body: str = Field(min_length=3, max_length=5000)
    recipient_id: int | None = None
    send_to_all: bool = False


class ClassIn(BaseModel):
    program_id: int
    name: str = Field(min_length=2, max_length=120)
    schedule: str = ""
    teacher: str = ""
    default_minutes: int = Field(default=60, ge=1, le=600)
    location: str = ""
    notes: str = ""


class ScheduleImportOptions(BaseModel):
    program_id: int
    activity_start: str = Field(min_length=8, max_length=20)
    activity_end: str = ""
    teacher: str = ""
    location: str = ""
    period_label: str = ""
    academic_year: int = Field(default=datetime.now().year, ge=2020, le=2100)
    include_only_confirmed: bool = True


class ClassStudentIn(StudentIn):
    user_id: int | None = None


class ClassSessionIn(BaseModel):
    session_date: str = Field(min_length=8, max_length=20)
    notes: str = ""


class AttendanceItemIn(BaseModel):
    user_id: int
    status: str = Field(pattern="^(present|absent|justified)$")
    minutes: int = Field(default=0, ge=0, le=600)
    notes: str = ""


class AttendanceBulkIn(BaseModel):
    records: list[AttendanceItemIn]


class AttendanceHistoryIn(BaseModel):
    user_id: int
    start_date: str = Field(min_length=8, max_length=20)
    end_date: str = ""
    status: str = Field(pattern="^(present|absent|justified)$")
    notes: str = ""


class AcademicReportQuery(BaseModel):
    program_id: int | None = None
    period_label: str = ""
    academic_year: int | None = None


def current_user(session: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> dict[str, Any]:
    if not session:
        raise HTTPException(401, "Login necessario.")
    token_hash = hashlib.sha256(session.encode()).hexdigest()
    with connect() as db:
        rec = db.execute(
            """
            SELECT u.* FROM sessions s
            JOIN users u ON u.id = s.user_id
            WHERE s.token_hash = ? AND s.expires_at > strftime('%s','now') AND u.active = 1
            """,
            (token_hash,),
        ).fetchone()
    if not rec:
        raise HTTPException(401, "Sessao invalida ou expirada.")
    return dict(rec)


def require_admin(user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    if user["role"] != "admin":
        raise HTTPException(403, "Acesso restrito aos estagiários/admin.")
    return user


def public_user(user: dict[str, Any]) -> dict[str, Any]:
    return {k: user[k] for k in user.keys() if k not in {"password_hash", "qr_secret"}}


def get_student_or_404(db: sqlite3.Connection, user_id: int) -> sqlite3.Row:
    row = db.execute("SELECT * FROM users WHERE id = ? AND role = 'student' AND active = 1", (user_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Aluno não encontrado.")
    return row


def toggle_attendance(db: sqlite3.Connection, student_id: int, admin_id: int, source: str) -> tuple[str, dict[str, Any]]:
    open_row = db.execute(
        "SELECT * FROM attendance WHERE user_id=? AND check_out IS NULL ORDER BY check_in DESC LIMIT 1",
        (student_id,),
    ).fetchone()
    stamp = now_iso()
    if open_row:
        start = datetime.fromisoformat(open_row["check_in"])
        end = datetime.fromisoformat(stamp)
        minutes = max(0, int((end - start).total_seconds() // 60))
        db.execute(
            "UPDATE attendance SET check_out=?, minutes=?, registered_by=?, source=? WHERE id=?",
            (stamp, minutes, admin_id, source, open_row["id"]),
        )
        return "saída", rowdict(db.execute("SELECT * FROM attendance WHERE id=?", (open_row["id"],)).fetchone())
    cur = db.execute(
        "INSERT INTO attendance (user_id, check_in, registered_by, source) VALUES (?, ?, ?, ?)",
        (student_id, stamp, admin_id, source),
    )
    return "entrada", rowdict(db.execute("SELECT * FROM attendance WHERE id=?", (cur.lastrowid,)).fetchone())


def calculate_result(data: EvaluationIn, goal: str) -> tuple[float, str, str]:
    bmi = None
    score = 100.0
    if data.weight is not None and data.height_cm is not None:
        height_m = data.height_cm / 100
        bmi = data.weight / (height_m * height_m)
        if bmi < 18.5 or bmi >= 30:
            score -= 18
        elif bmi >= 25:
            score -= 9
    if data.body_fat is not None:
        score -= max(0, data.body_fat - 25) * 0.9
    if data.visceral_fat is not None and data.visceral_fat >= 12:
        score -= 10
    if data.vo2max is not None:
        score += min(10, max(-15, (data.vo2max - 35) * 0.8))
    if data.resting_hr is not None and data.resting_hr > 85:
        score -= 8
    if data.systolic and data.diastolic and (data.systolic >= 140 or data.diastolic >= 90):
        score -= 15
    if data.flexibility_cm is not None and data.flexibility_cm < 20:
        score -= 5
    score = round(max(0, min(100, score)), 1)
    risk = "baixo" if score >= 78 else "moderado" if score >= 55 else "alto"
    focus = {
        "emagrecimento": "priorizar cardio progressivo, treino de força multiarticular e controle de intensidade.",
        "hipertrofia": "priorizar musculação com progressão semanal, descanso e técnica.",
        "condicionamento": "priorizar capacidade aeróbica, esteira intervalada leve e resistência muscular.",
        "saude": "combinar esteira, aparelhos de força, mobilidade e consistência semanal.",
    }.get(goal, "combinar esteira, aparelhos de força, mobilidade e consistência semanal.")
    recs = [
        f"IMC estimado: {bmi:.1f}. Objetivo declarado: {goal}." if bmi is not None else f"Avaliação parcial registrada. Objetivo declarado: {goal}.",
        f"Foco do mês: {focus}",
        "Semana 1: adaptação, cargas confortáveis, 2-3 séries de 12-15 repetições.",
        "Semana 2: aumentar levemente volume ou carga mantendo execução segura.",
        "Semana 3: consolidar progressão, incluir intervalos curtos na esteira se liberado.",
        "Semana 4: reduzir fadiga, reavaliar medidas e ajustar o próximo ciclo.",
    ]
    if risk == "alto":
        recs.append("Recomendado acompanhamento próximo do estagiário/profissional antes de intensificar.")
    return score, risk, "\n".join(recs)


@app.post("/api/auth/login")
def login(data: LoginIn, response: Response) -> dict[str, Any]:
    with connect() as db:
        user = db.execute("SELECT * FROM users WHERE email = ? AND active = 1", (normalize_email(data.email),)).fetchone()
        if not user or not verify_password(data.password, user["password_hash"]):
            raise HTTPException(401, "E-mail ou senha inválidos.")
        token = secrets.token_urlsafe(48)
        db.execute(
            "INSERT INTO sessions (token_hash, user_id, expires_at, created_at) VALUES (?, ?, strftime('%s','now') + ?, ?)",
            (hashlib.sha256(token.encode()).hexdigest(), user["id"], SESSION_SECONDS, now_iso()),
        )
    response.set_cookie(COOKIE_NAME, token, max_age=SESSION_SECONDS, httponly=True, secure=COOKIE_SECURE, samesite="lax")
    return {"user": public_user(dict(user))}


@app.post("/api/auth/logout")
def logout(response: Response, session: str | None = Cookie(default=None, alias=COOKIE_NAME)) -> dict[str, str]:
    if session:
        with connect() as db:
            db.execute("DELETE FROM sessions WHERE token_hash = ?", (hashlib.sha256(session.encode()).hexdigest(),))
    response.delete_cookie(COOKIE_NAME)
    return {"status": "ok"}


@app.post("/api/auth/register")
def public_register(data: PublicRegisterIn) -> dict[str, str]:
    with connect() as db:
        try:
            db.execute(
                """
                INSERT INTO users
                (name,email,password_hash,role,cpf,registration,birth_date,phone,goal,availability_days,weekly_minutes,notes,qr_secret,active,created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    data.name.strip(),
                    normalize_email(data.email),
                    hash_password(data.password),
                    "student",
                    data.cpf,
                    data.registration,
                    data.birth_date,
                    data.phone,
                    data.goal,
                    data.availability_days,
                    data.weekly_minutes,
                    "Cadastro solicitado pelo aluno. Aguardando aprovação.",
                    secrets.token_urlsafe(32),
                    0,
                    now_iso(),
                ),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(409, "E-mail já cadastrado ou aguardando aprovação.")
    return {"status": "pending", "message": "Cadastro solicitado. Aguarde aprovação da equipe CAFIS."}


@app.get("/api/me")
def me(user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    return {"user": public_user(user)}


@app.get("/api/admin/overview")
def overview(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        total_students = db.execute("SELECT count(*) c FROM users WHERE role='student' AND active=1").fetchone()["c"]
        open_att = db.execute("SELECT count(*) c FROM attendance WHERE check_out IS NULL").fetchone()["c"]
        minutes = db.execute("SELECT coalesce(sum(minutes),0) c FROM attendance").fetchone()["c"]
        evaluations = db.execute("SELECT count(*) c FROM evaluations").fetchone()["c"]
    return {"students": total_students, "inside_now": open_att, "total_hours": round(minutes / 60, 1), "evaluations": evaluations}


@app.get("/api/students")
def list_students(search: str = "", _: dict[str, Any] = Depends(require_admin)) -> list[dict[str, Any]]:
    with connect() as db:
        term = f"%{search.strip()}%" if search.strip() else "%"
        rows = db.execute(
            """
            SELECT u.*, (SELECT max(created_at) FROM evaluations e WHERE e.user_id=u.id) last_evaluation,
                   (SELECT coalesce(sum(minutes),0) FROM attendance a WHERE a.user_id=u.id) total_minutes,
                   EXISTS(SELECT 1 FROM attendance a2 WHERE a2.user_id=u.id AND a2.check_out IS NULL) inside_now,
                   (SELECT count(*) FROM class_attendance ca WHERE ca.user_id=u.id AND ca.status='absent') total_absences,
                   (SELECT count(*) FROM class_attendance ca WHERE ca.user_id=u.id AND ca.status='present') total_class_presences
            FROM users u
            WHERE u.role='student' AND u.active=1
              AND (u.name LIKE ? OR u.email LIKE ? OR coalesce(u.registration,'') LIKE ?)
            ORDER BY u.name
            """,
            (term, term, term),
        ).fetchall()
    return [
        public_user(dict(r)) | {
            "last_evaluation": r["last_evaluation"],
            "total_hours": round(r["total_minutes"] / 60, 1),
            "inside_now": bool(r["inside_now"]),
            "total_absences": r["total_absences"] or 0,
            "total_class_presences": r["total_class_presences"] or 0,
        }
        for r in rows
    ]


@app.get("/api/admin/pending-students")
def pending_students(_: dict[str, Any] = Depends(require_admin)) -> list[dict[str, Any]]:
    with connect() as db:
        rows = db.execute("SELECT * FROM users WHERE role='student' AND active=0 ORDER BY created_at DESC").fetchall()
    return [public_user(dict(r)) for r in rows]


@app.post("/api/admin/pending-students/{student_id}/approve")
def approve_student(student_id: int, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        row = db.execute("SELECT * FROM users WHERE id=? AND role='student' AND active=0", (student_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Solicitação não encontrada.")
        db.execute("UPDATE users SET active=1, notes=? WHERE id=?", ("Cadastro aprovado pela equipe CAFIS.", student_id))
        user = rowdict(db.execute("SELECT * FROM users WHERE id=?", (student_id,)).fetchone())
    return {"student": public_user(user)}


@app.delete("/api/admin/pending-students/{student_id}")
def reject_student(student_id: int, _: dict[str, Any] = Depends(require_admin)) -> dict[str, str]:
    with connect() as db:
        row = db.execute("SELECT id FROM users WHERE id=? AND role='student' AND active=0", (student_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Solicitação não encontrada.")
        db.execute("DELETE FROM users WHERE id=?", (student_id,))
    return {"status": "rejected"}


@app.post("/api/students")
def create_student(data: StudentIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    password = data.password or (data.cpf[-4:] if len(data.cpf) >= 4 else secrets.token_urlsafe(6))
    with connect() as db:
        try:
            cur = db.execute(
                """
                INSERT INTO users
                (name,email,password_hash,role,cpf,registration,birth_date,phone,goal,availability_days,weekly_minutes,notes,qr_secret,created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    data.name.strip(),
                    normalize_email(data.email),
                    hash_password(password),
                    "student",
                    data.cpf,
                    data.registration,
                    data.birth_date,
                    data.phone,
                    data.goal,
                    data.availability_days,
                    data.weekly_minutes,
                    data.notes,
                    secrets.token_urlsafe(32),
                    now_iso(),
                ),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(409, "E-mail já cadastrado.")
        user = rowdict(db.execute("SELECT * FROM users WHERE id=?", (cur.lastrowid,)).fetchone())
    return {"student": public_user(user), "initial_password": password}


@app.put("/api/students/{student_id}")
def update_student(student_id: int, data: StudentIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        get_student_or_404(db, student_id)
        db.execute(
            """
            UPDATE users SET name=?, email=?, cpf=?, registration=?, birth_date=?, phone=?, goal=?,
            availability_days=?, weekly_minutes=?, notes=? WHERE id=?
            """,
            (data.name, normalize_email(data.email), data.cpf, data.registration, data.birth_date, data.phone,
             data.goal, data.availability_days, data.weekly_minutes, data.notes, student_id),
        )
        if data.password:
            db.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password(data.password), student_id))
        user = rowdict(db.execute("SELECT * FROM users WHERE id=?", (student_id,)).fetchone())
    return {"student": public_user(user)}


@app.get("/api/students/{student_id}")
def student_detail(student_id: int, admin: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    if admin["role"] != "admin" and admin["id"] != student_id:
        raise HTTPException(403, "Sem permissao.")
    with connect() as db:
        student = rowdict(get_student_or_404(db, student_id))
        evals = [dict(r) for r in db.execute("SELECT * FROM evaluations WHERE user_id=? ORDER BY created_at", (student_id,)).fetchall()]
        attendance = [dict(r) for r in db.execute("SELECT * FROM attendance WHERE user_id=? ORDER BY check_in DESC LIMIT 90", (student_id,)).fetchall()]
        plans = [dict(r) for r in db.execute("SELECT * FROM workout_plans WHERE user_id=? ORDER BY created_at DESC LIMIT 8", (student_id,)).fetchall()]
        class_summary = [dict(r) for r in db.execute(
            """
            SELECT c.id class_id, c.name class_name, p.name program_name, c.period_label, c.academic_year,
                   SUM(CASE WHEN a.status='present' THEN 1 ELSE 0 END) presents,
                   SUM(CASE WHEN a.status='absent' THEN 1 ELSE 0 END) absences,
                   SUM(CASE WHEN a.status='justified' THEN 1 ELSE 0 END) justified,
                   SUM(CASE WHEN a.status='present' THEN a.minutes ELSE 0 END) minutes
            FROM enrollments e
            JOIN class_groups c ON c.id=e.class_id
            JOIN programs p ON p.id=c.program_id
            LEFT JOIN class_sessions s ON s.class_id=c.id
            LEFT JOIN class_attendance a ON a.session_id=s.id AND a.user_id=e.user_id
            WHERE e.user_id=?
            GROUP BY c.id, c.name, p.name, c.period_label, c.academic_year
            ORDER BY p.name, c.weekday, c.start_time, c.name
            """,
            (student_id,),
        ).fetchall()]
    return {"student": public_user(student), "evaluations": evals, "attendance": attendance, "plans": plans, "class_summary": class_summary}


@app.get("/api/my/dashboard")
def my_dashboard(user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    return student_detail(user["id"], user)


@app.post("/api/evaluations")
def create_evaluation(data: EvaluationIn, user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    student_id = data.user_id if user["role"] == "admin" else user["id"]
    if not student_id:
        raise HTTPException(400, "Aluno obrigatório.")
    with connect() as db:
        student = get_student_or_404(db, student_id)
        has_metric = any(
            getattr(data, field) is not None
            for field in (
                "weight", "height_cm", "body_fat", "body_water", "muscle_mass", "bmr", "metabolic_age",
                "bone_mass", "visceral_fat", "waist_cm", "hip_cm", "flexibility_cm", "abdominal_reps",
                "pushup_reps", "resting_hr", "post_hr", "recovery_hr_5min", "cooper_km", "vo2max",
                "systolic", "diastolic",
            )
        )
        if not has_metric:
            raise HTTPException(400, "Preencha pelo menos um resultado da avaliação.")
        score, risk, recs = calculate_result(data, student["goal"])
        cur = db.execute(
            """
            INSERT INTO evaluations
            (user_id,evaluation_type,created_at,weight,height_cm,body_fat,body_water,muscle_mass,bmr,metabolic_age,bone_mass,visceral_fat,
             waist_cm,hip_cm,flexibility_cm,abdominal_reps,pushup_reps,resting_hr,post_hr,recovery_hr_5min,cooper_km,
             vo2max,systolic,diastolic,qualia_score,risk_level,recommendations,notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (student_id, data.evaluation_type, now_iso(), data.weight, data.height_cm, data.body_fat, data.body_water, data.muscle_mass,
             data.bmr, data.metabolic_age, data.bone_mass, data.visceral_fat, data.waist_cm, data.hip_cm,
             data.flexibility_cm, data.abdominal_reps, data.pushup_reps, data.resting_hr, data.post_hr,
             data.recovery_hr_5min, data.cooper_km, data.vo2max, data.systolic, data.diastolic, score, risk, recs, data.notes),
        )
        record = rowdict(db.execute("SELECT * FROM evaluations WHERE id=?", (cur.lastrowid,)).fetchone())
    return {"evaluation": record}


@app.get("/api/my/qr")
def my_qr(user: dict[str, Any] = Depends(current_user)) -> dict[str, str]:
    payload = f"CAFIS:{user['id']}:{user['qr_secret']}"
    img = qrcode.make(payload)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return {"payload": payload, "png": "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()}


@app.post("/api/attendance/scan")
def scan_attendance(data: AttendanceScanIn, admin: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    parts = data.qr_payload.strip().split(":")
    if len(parts) != 3 or parts[0] != "CAFIS":
        raise HTTPException(400, "QR code inválido.")
    try:
        student_id = int(parts[1])
    except ValueError:
        raise HTTPException(400, "QR code inválido.")
    with connect() as db:
        student = get_student_or_404(db, student_id)
        if not hmac.compare_digest(parts[2], student["qr_secret"]):
            raise HTTPException(403, "QR code não pertence ao aluno.")
        action, rec = toggle_attendance(db, student_id, admin["id"], "qr")
    return {"action": action, "student": public_user(dict(student)), "attendance": rec}


@app.post("/api/attendance/manual")
def manual_attendance(data: AttendanceManualIn, admin: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        student = get_student_or_404(db, data.user_id)
        action, rec = toggle_attendance(db, data.user_id, admin["id"], "manual")
    return {"action": action, "student": public_user(dict(student)), "attendance": rec}


@app.get("/api/equipment")
def list_equipment(_: dict[str, Any] = Depends(current_user)) -> list[dict[str, Any]]:
    with connect() as db:
        return [dict(r) for r in db.execute("SELECT * FROM equipment ORDER BY category, name").fetchall()]


@app.post("/api/equipment")
def add_equipment(data: EquipmentIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        cur = db.execute(
            """
            INSERT INTO equipment
            (name,category,status,quantity,location,photo_data_url,maintenance_notes,created_at)
            VALUES (?,?,?,?,?,?,?,?)
            """,
            (
                data.name,
                data.category,
                data.status,
                data.quantity,
                data.location,
                data.photo_data_url,
                data.maintenance_notes,
                now_iso(),
            ),
        )
        return {"equipment": rowdict(db.execute("SELECT * FROM equipment WHERE id=?", (cur.lastrowid,)).fetchone())}


@app.put("/api/equipment/{equipment_id}")
def update_equipment(equipment_id: int, data: EquipmentIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        row = db.execute("SELECT id FROM equipment WHERE id=?", (equipment_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Equipamento não encontrado.")
        db.execute(
            """
            UPDATE equipment SET name=?, category=?, status=?, quantity=?, location=?,
            photo_data_url=?, maintenance_notes=? WHERE id=?
            """,
            (
                data.name,
                data.category,
                data.status,
                data.quantity,
                data.location,
                data.photo_data_url,
                data.maintenance_notes,
                equipment_id,
            ),
        )
        return {"equipment": rowdict(db.execute("SELECT * FROM equipment WHERE id=?", (equipment_id,)).fetchone())}


@app.post("/api/workout-plans/{student_id}")
def save_plan(student_id: int, body: dict[str, Any], admin: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    text = str(body.get("plan_text", "")).strip()
    if len(text) < 10:
        raise HTTPException(400, "Plano muito curto.")
    with connect() as db:
        get_student_or_404(db, student_id)
        cur = db.execute(
            "INSERT INTO workout_plans (user_id,month,week,plan_text,created_by,created_at) VALUES (?,?,?,?,?,?)",
            (student_id, str(body.get("month") or now_iso()[:7]), int(body.get("week") or 1), text, admin["id"], now_iso()),
        )
        return {"plan": rowdict(db.execute("SELECT * FROM workout_plans WHERE id=?", (cur.lastrowid,)).fetchone())}


def send_mail(to_email: str, subject: str, body: str) -> str:
    host = os.getenv("CAFIS_SMTP_HOST", "")
    if not host:
        return "registrado_sem_smtp"
    msg = EmailMessage()
    msg["From"] = os.getenv("CAFIS_SMTP_FROM", os.getenv("CAFIS_SMTP_USER", "no-reply@cafis.local"))
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)
    with smtplib.SMTP(host, int(os.getenv("CAFIS_SMTP_PORT", "587"))) as smtp:
        smtp.starttls()
        user = os.getenv("CAFIS_SMTP_USER", "")
        password = os.getenv("CAFIS_SMTP_PASSWORD", "")
        if user:
            smtp.login(user, password)
        smtp.send_message(msg)
    return "enviado"


@app.post("/api/messages")
def send_message(data: MessageIn, admin: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        targets = []
        if data.send_to_all:
            targets = db.execute("SELECT id,email FROM users WHERE role='student' AND active=1").fetchall()
        elif data.recipient_id:
            targets = [get_student_or_404(db, data.recipient_id)]
        if not targets:
            raise HTTPException(400, "Selecione um aluno ou envio para todos.")
        results = []
        for target in targets:
            status = send_mail(target["email"], data.subject, data.body)
            db.execute(
                "INSERT INTO messages (sender_id,recipient_id,subject,body,status,created_at) VALUES (?,?,?,?,?,?)",
                (admin["id"], target["id"], data.subject, data.body, status, now_iso()),
            )
            results.append({"email": target["email"], "status": status})
    return {"sent": results}


def get_class_or_404(db: sqlite3.Connection, class_id: int) -> sqlite3.Row:
    row = db.execute(
        """
        SELECT c.*, p.name program_name, p.slug program_slug
        FROM class_groups c JOIN programs p ON p.id = c.program_id
        WHERE c.id=? AND c.active=1
        """,
        (class_id,),
    ).fetchone()
    if not row:
        raise HTTPException(404, "Turma não encontrada.")
    return row


def upsert_student_from_payload(db: sqlite3.Connection, data: dict[str, Any]) -> tuple[int, str | None]:
    email = normalize_email(str(data.get("email") or ""))
    if not email:
        raise HTTPException(400, "E-mail do aluno é obrigatório.")
    existing = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
    if existing:
        return existing["id"], None
    cpf = str(data.get("cpf") or "")
    password = str(data.get("password") or (cpf[-4:] if len(cpf) >= 4 else secrets.token_urlsafe(6)))
    cur = db.execute(
        """
        INSERT INTO users
        (name,email,password_hash,role,cpf,registration,birth_date,phone,goal,availability_days,weekly_minutes,notes,qr_secret,active,created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            str(data.get("name") or data.get("nome") or "").strip(),
            email,
            hash_password(password),
            "student",
            cpf,
            str(data.get("registration") or data.get("matricula") or data.get("ra") or ""),
            str(data.get("birth_date") or data.get("nascimento") or ""),
            str(data.get("phone") or data.get("telefone") or ""),
            str(data.get("goal") or "saude"),
            str(data.get("availability_days") or data.get("dias") or ""),
            int(data.get("weekly_minutes") or 180),
            str(data.get("notes") or ""),
            secrets.token_urlsafe(32),
            1,
            now_iso(),
        ),
    )
    return int(cur.lastrowid), password


def build_class_name(program_name: str, weekday: str, start_time: str, end_time: str) -> str:
    return f"{program_name} - {weekday} {start_time}-{end_time}".strip()


def generate_class_sessions(db: sqlite3.Connection, class_id: int, start_date_text: str, end_date_text: str, weekday_text: str) -> int:
    if not start_date_text or not end_date_text or not weekday_text:
        return 0
    start_date = date.fromisoformat(start_date_text)
    end_date = date.fromisoformat(end_date_text)
    weekday_key = normalize_text(weekday_text)
    weekday_index = WEEKDAY_INDEX.get(weekday_key)
    if weekday_index is None:
        return 0
    created = 0
    for current in daterange(start_date, end_date):
        if current.weekday() != weekday_index:
            continue
        before = db.total_changes
        db.execute(
            "INSERT OR IGNORE INTO class_sessions (class_id,session_date,notes,created_at) VALUES (?,?,?,?)",
            (class_id, current.isoformat(), "", now_iso()),
        )
        if db.total_changes > before:
            session_id = db.execute(
                "SELECT id FROM class_sessions WHERE class_id=? AND session_date=?",
                (class_id, current.isoformat()),
            ).fetchone()["id"]
            students = db.execute("SELECT user_id FROM enrollments WHERE class_id=?", (class_id,)).fetchall()
            for student in students:
                db.execute(
                    """
                    INSERT OR IGNORE INTO class_attendance
                    (session_id,user_id,status,minutes,registered_at)
                    VALUES (?,?,?,?,?)
                    """,
                    (session_id, student["user_id"], "absent", 0, now_iso()),
                )
            created += 1
    return created


def parse_workbook_groups(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    groups: list[dict[str, Any]] = []
    valid_sheets = {normalize_text(item) for item in IMPORT_SHEETS}
    for sheet_name in workbook.sheetnames:
        normalized_sheet = normalize_text(sheet_name)
        if normalized_sheet not in valid_sheets:
            continue
        weekday_label = canonical_weekday_label(sheet_name)
        ws = workbook[sheet_name]
        current_group: dict[str, Any] | None = None
        for row in ws.iter_rows(values_only=True):
            first = clean_label(row[0] if row else "")
            normalized_first = normalize_text(first)
            if first and ":" in first and "feira" in normalized_first:
                start_time, end_time = parse_time_range(first)
                current_group = {
                    "weekday": weekday_label,
                    "title": first,
                    "start_time": start_time,
                    "end_time": end_time,
                    "columns": {},
                    "students": [],
                }
                groups.append(current_group)
                continue
            if not current_group:
                continue
            if normalized_first.startswith("cod. evento"):
                columns = {}
                for index, header in enumerate(row):
                    normalized_header = normalize_text(header)
                    if normalized_header:
                        columns[normalized_header] = index
                current_group["columns"] = columns
                continue
            if normalized_first.startswith("confirmado(s)"):
                current_group = None
                continue
            columns = current_group.get("columns", {})
            def col(*names: str):
                for name in names:
                    idx = columns.get(normalize_text(name))
                    if idx is not None and idx < len(row):
                        return row[idx]
                return None
            code = col("Cod. Evento") if columns else (row[0] if len(row) > 0 else None)
            registration = col("Nº Inscrição", "N� Inscri��o") if columns else (row[1] if len(row) > 1 else None)
            name = clean_label(col("Nome da Pessoa") if columns else (row[2] if len(row) > 2 else ""))
            birth_value = col("Data de Nascimento") if columns else (row[3] if len(row) > 3 else None)
            sex = clean_label(col("Sexo") if columns else (row[4] if len(row) > 4 else ""))
            phone_home = clean_label(col("Tel. Residencial") if columns else (row[5] if len(row) > 5 else ""))
            phone_mobile = clean_label(col("Tel. Celular") if columns else (row[6] if len(row) > 6 else ""))
            email = clean_label(col("E-mail") if columns else (row[7] if len(row) > 7 else ""))
            status = clean_label(col("Situação", "Situa��o") if columns else (row[8] if len(row) > 8 else ""))
            if not name:
                continue
            normalized_status = normalize_text(status)
            if normalized_status and any(flag in normalized_status for flag in ("cancelado", "desistente")):
                continue
            current_group["students"].append(
                {
                    "event_code": only_digits(code),
                    "registration": only_digits(registration) or clean_label(registration),
                    "name": name,
                    "birth_date": as_iso_date(birth_value),
                    "sex": sex,
                    "phone": phone_mobile or phone_home,
                    "email": email,
                    "status": status or "Confirmado",
                }
            )
    return [group for group in groups if group["students"]]


def import_schedule_workbook(
    db: sqlite3.Connection,
    *,
    program_id: int,
    file_name: str,
    file_bytes: bytes,
    activity_start: str,
    activity_end: str,
    teacher: str,
    location: str,
    period_label: str,
    academic_year: int,
) -> dict[str, Any]:
    program = db.execute("SELECT * FROM programs WHERE id=?", (program_id,)).fetchone()
    if not program:
        raise HTTPException(404, "Projeto não encontrado.")
    groups = parse_workbook_groups(file_bytes)
    if not groups:
        raise HTTPException(400, "Não encontrei turmas válidas nas abas Segunda a Sexta da planilha.")
    created_classes = 0
    created_students = 0
    enrollments = 0
    sessions_created = 0
    imported_classes: list[dict[str, Any]] = []
    program_name = program["name"]
    effective_end = activity_end or f"{academic_year}-12-31"
    for group in groups:
        class_name = build_class_name(program_name, group["weekday"], group["start_time"], group["end_time"])
        existing = db.execute(
            """
            SELECT id FROM class_groups
            WHERE program_id=? AND name=? AND weekday=? AND start_time=? AND end_time=? AND academic_year=? AND active=1
            """,
            (program_id, class_name, group["weekday"], group["start_time"], group["end_time"], academic_year),
        ).fetchone()
        if existing:
            class_id = existing["id"]
        else:
            cur = db.execute(
                """
                INSERT INTO class_groups
                (program_id,name,schedule,teacher,default_minutes,location,weekday,start_time,end_time,activity_start,activity_end,academic_year,period_label,source_file,event_code,notes,created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    program_id,
                    class_name,
                    f"{group['weekday']}, {group['start_time']} às {group['end_time']}",
                    teacher,
                    50,
                    location,
                    group["weekday"],
                    group["start_time"],
                    group["end_time"],
                    activity_start,
                    effective_end,
                    academic_year,
                    period_label,
                    file_name,
                    group["students"][0].get("event_code", ""),
                    group["title"],
                    now_iso(),
                ),
            )
            class_id = int(cur.lastrowid)
            created_classes += 1
        imported_classes.append({"class_id": class_id, "name": class_name, "weekday": group["weekday"], "students": len(group["students"])})
        for student in group["students"]:
            fallback_email = ""
            if not student["email"]:
                reg = student["registration"] or only_digits(student["birth_date"]) or secrets.token_hex(4)
                fallback_email = f"aluno{reg}@cafis.local"
            payload = {
                "name": student["name"],
                "email": student["email"] or fallback_email,
                "registration": student["registration"],
                "birth_date": student["birth_date"],
                "phone": student["phone"],
                "notes": f"Importado de {file_name} | Status: {student['status']}",
                "weekly_minutes": 50,
                "availability_days": group["weekday"],
            }
            user_id, initial_password = upsert_student_from_payload(db, payload)
            if initial_password:
                created_students += 1
            before = db.total_changes
            db.execute(
                "INSERT OR IGNORE INTO enrollments (class_id,user_id,created_at) VALUES (?,?,?)",
                (class_id, user_id, now_iso()),
            )
            if db.total_changes > before:
                enrollments += 1
        sessions_created += generate_class_sessions(db, class_id, activity_start, effective_end, group["weekday"])
    return {
        "classes_created": created_classes,
        "students_created": created_students,
        "enrollments_created": enrollments,
        "sessions_created": sessions_created,
        "classes": imported_classes,
    }


def class_report_summary(db: sqlite3.Connection, class_id: int) -> list[dict[str, Any]]:
    rows = db.execute(
        """
        SELECT u.id user_id, u.name, u.email, u.registration,
               SUM(CASE WHEN a.status='present' THEN 1 ELSE 0 END) presents,
               SUM(CASE WHEN a.status='absent' THEN 1 ELSE 0 END) absences,
               SUM(CASE WHEN a.status='justified' THEN 1 ELSE 0 END) justified,
               SUM(CASE WHEN a.status='present' THEN a.minutes ELSE 0 END) minutes
        FROM enrollments e
        JOIN users u ON u.id = e.user_id
        LEFT JOIN class_sessions s ON s.class_id = e.class_id
        LEFT JOIN class_attendance a ON a.session_id = s.id AND a.user_id = u.id
        WHERE e.class_id = ?
        GROUP BY u.id, u.name, u.email, u.registration
        ORDER BY u.name
        """,
        (class_id,),
    ).fetchall()
    return [dict(row) for row in rows]


@app.get("/api/programs")
def programs(_: dict[str, Any] = Depends(require_admin)) -> list[dict[str, Any]]:
    with connect() as db:
        rows = db.execute(
            """
            SELECT p.*,
                   (SELECT count(*) FROM class_groups c WHERE c.program_id=p.id AND c.active=1) class_count
            FROM programs p ORDER BY p.name
            """
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/classes")
def classes(
    program_id: int | None = None,
    period_label: str = "",
    academic_year: int | None = None,
    _: dict[str, Any] = Depends(require_admin),
) -> list[dict[str, Any]]:
    with connect() as db:
        where = ["c.active=1"]
        params: list[Any] = []
        if program_id:
            where.append("c.program_id=?")
            params.append(program_id)
        if period_label.strip():
            where.append("lower(c.period_label)=lower(?)")
            params.append(period_label.strip())
        if academic_year:
            where.append("c.academic_year=?")
            params.append(academic_year)
        rows = db.execute(
            f"""
            SELECT c.*, p.name program_name,
                   (SELECT count(*) FROM enrollments e WHERE e.class_id=c.id) student_count,
                   (SELECT count(*) FROM class_sessions s WHERE s.class_id=c.id) session_count
            FROM class_groups c JOIN programs p ON p.id=c.program_id
            WHERE {' AND '.join(where)} ORDER BY p.name, c.weekday, c.start_time, c.name
            """,
            tuple(params),
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/programs/{program_id}/schedule-import")
async def schedule_import(
    program_id: int,
    activity_start: str = Form(...),
    activity_end: str = Form(""),
    teacher: str = Form(""),
    location: str = Form(""),
    period_label: str = Form(""),
    academic_year: int = Form(default=datetime.now().year),
    file: UploadFile = File(...),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Arquivo vazio.")
    start_iso = as_iso_date(activity_start)
    if not start_iso:
        raise HTTPException(400, "Informe a data de início da atividade.")
    end_iso = as_iso_date(activity_end) if activity_end else f"{academic_year}-12-31"
    with connect() as db:
        result = import_schedule_workbook(
            db,
            program_id=program_id,
            file_name=file.filename or "planilha.xlsx",
            file_bytes=raw,
            activity_start=start_iso,
            activity_end=end_iso,
            teacher=teacher,
            location=location,
            period_label=period_label or f"{academic_year}",
            academic_year=academic_year,
        )
    return result


@app.post("/api/classes")
def create_class(data: ClassIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        if not db.execute("SELECT id FROM programs WHERE id=?", (data.program_id,)).fetchone():
            raise HTTPException(404, "Projeto não encontrado.")
        cur = db.execute(
            """
            INSERT INTO class_groups
            (program_id,name,schedule,teacher,default_minutes,location,notes,created_at)
            VALUES (?,?,?,?,?,?,?,?)
            """,
            (data.program_id, data.name, data.schedule, data.teacher, data.default_minutes, data.location, data.notes, now_iso()),
        )
        row = rowdict(db.execute("SELECT * FROM class_groups WHERE id=?", (cur.lastrowid,)).fetchone())
    return {"class": row}


@app.put("/api/classes/{class_id}")
def update_class(class_id: int, data: ClassIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        get_class_or_404(db, class_id)
        db.execute(
            """
            UPDATE class_groups SET program_id=?, name=?, schedule=?, teacher=?,
            default_minutes=?, location=?, notes=? WHERE id=?
            """,
            (data.program_id, data.name, data.schedule, data.teacher, data.default_minutes, data.location, data.notes, class_id),
        )
        row = rowdict(db.execute("SELECT * FROM class_groups WHERE id=?", (class_id,)).fetchone())
    return {"class": row}


@app.get("/api/classes/{class_id}/roster")
def class_roster(class_id: int, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        class_row = rowdict(get_class_or_404(db, class_id))
        students = [
            public_user(dict(r)) for r in db.execute(
                """
                SELECT u.* FROM enrollments e JOIN users u ON u.id=e.user_id
                WHERE e.class_id=? AND u.active=1 ORDER BY u.name
                """,
                (class_id,),
            ).fetchall()
        ]
        sessions = [
            dict(r) for r in db.execute(
                "SELECT * FROM class_sessions WHERE class_id=? ORDER BY session_date DESC LIMIT 60",
                (class_id,),
            ).fetchall()
        ]
        report = class_report_summary(db, class_id)
    return {"class": class_row, "students": students, "sessions": sessions, "report": report}


@app.post("/api/classes/{class_id}/students")
def add_class_student(class_id: int, data: ClassStudentIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        get_class_or_404(db, class_id)
        if data.user_id:
            user_id = data.user_id
            initial_password = None
            get_student_or_404(db, user_id)
        else:
            user_id, initial_password = upsert_student_from_payload(db, data.model_dump())
        db.execute(
            "INSERT OR IGNORE INTO enrollments (class_id,user_id,created_at) VALUES (?,?,?)",
            (class_id, user_id, now_iso()),
        )
        existing_sessions = db.execute("SELECT id FROM class_sessions WHERE class_id=?", (class_id,)).fetchall()
        for session in existing_sessions:
            db.execute(
                """
                INSERT OR IGNORE INTO class_attendance
                (session_id,user_id,status,minutes,registered_at)
                VALUES (?,?,?,?,?)
                """,
                (session["id"], user_id, "absent", 0, now_iso()),
            )
        student = rowdict(db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone())
    return {"student": public_user(student), "initial_password": initial_password}


def read_students_table(upload: UploadFile, raw: bytes) -> list[dict[str, Any]]:
    suffix = Path(upload.filename or "").suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise HTTPException(500, "Dependência openpyxl não instalada.") from exc
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(v or "").strip().lower() for v in rows[0]]
        return [
            {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
            for row in rows[1:]
            if any(cell not in (None, "") for cell in row)
        ]
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


@app.post("/api/classes/{class_id}/import")
async def import_class_students(
    class_id: int,
    file: UploadFile = File(...),
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    raw = await file.read()
    rows = read_students_table(file, raw)
    created = 0
    enrolled = 0
    failures: list[dict[str, Any]] = []
    with connect() as db:
        get_class_or_404(db, class_id)
        for idx, row in enumerate(rows, start=2):
            try:
                normalized = {
                    "name": row.get("nome") or row.get("name"),
                    "email": row.get("email") or row.get("e-mail"),
                    "cpf": row.get("cpf"),
                    "registration": row.get("matricula") or row.get("matrícula") or row.get("ra") or row.get("registration"),
                    "birth_date": row.get("nascimento") or row.get("birth_date"),
                    "phone": row.get("telefone") or row.get("phone"),
                    "password": row.get("senha") or row.get("password"),
                    "notes": row.get("observacoes") or row.get("observações") or "",
                }
                user_id, initial_password = upsert_student_from_payload(db, normalized)
                if initial_password:
                    created += 1
                before = db.total_changes
                db.execute(
                    "INSERT OR IGNORE INTO enrollments (class_id,user_id,created_at) VALUES (?,?,?)",
                    (class_id, user_id, now_iso()),
                )
                if db.total_changes > before:
                    enrolled += 1
                for session in db.execute("SELECT id FROM class_sessions WHERE class_id=?", (class_id,)).fetchall():
                    db.execute(
                        """
                        INSERT OR IGNORE INTO class_attendance
                        (session_id,user_id,status,minutes,registered_at)
                        VALUES (?,?,?,?,?)
                        """,
                        (session["id"], user_id, "absent", 0, now_iso()),
                    )
            except Exception as exc:
                failures.append({"row": idx, "error": str(exc)})
    return {"rows": len(rows), "created": created, "enrolled": enrolled, "failures": failures}


@app.post("/api/classes/{class_id}/sessions")
def create_class_session(class_id: int, data: ClassSessionIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        class_row = get_class_or_404(db, class_id)
        cur = db.execute(
            "INSERT OR IGNORE INTO class_sessions (class_id,session_date,notes,created_at) VALUES (?,?,?,?)",
            (class_id, data.session_date, data.notes, now_iso()),
        )
        session_id = cur.lastrowid or db.execute(
            "SELECT id FROM class_sessions WHERE class_id=? AND session_date=?",
            (class_id, data.session_date),
        ).fetchone()["id"]
        students = db.execute(
            "SELECT user_id FROM enrollments WHERE class_id=?",
            (class_id,),
        ).fetchall()
        for student in students:
            db.execute(
                """
                INSERT OR IGNORE INTO class_attendance
                (session_id,user_id,status,minutes,registered_at)
                VALUES (?,?,?,?,?)
                """,
                (session_id, student["user_id"], "absent", 0, now_iso()),
            )
        session = rowdict(db.execute("SELECT * FROM class_sessions WHERE id=?", (session_id,)).fetchone())
    return {"class": dict(class_row), "session": session}


@app.get("/api/sessions/{session_id}/attendance")
def get_session_attendance(session_id: int, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        session = rowdict(db.execute("SELECT * FROM class_sessions WHERE id=?", (session_id,)).fetchone())
        if not session:
            raise HTTPException(404, "Aula não encontrada.")
        class_row = rowdict(get_class_or_404(db, session["class_id"]))
        rows = db.execute(
            """
            SELECT u.id user_id, u.name, u.email, u.registration,
                   COALESCE(a.status, 'absent') status,
                   COALESCE(a.minutes, 0) minutes,
                   COALESCE(a.notes, '') notes
            FROM enrollments e
            JOIN users u ON u.id=e.user_id
            LEFT JOIN class_attendance a ON a.user_id=u.id AND a.session_id=?
            WHERE e.class_id=? AND u.active=1
            ORDER BY u.name
            """,
            (session_id, session["class_id"]),
        ).fetchall()
    return {"class": class_row, "session": session, "attendance": [dict(r) for r in rows]}


@app.get("/api/classes/{class_id}/report")
def class_report(class_id: int, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        class_row = rowdict(get_class_or_404(db, class_id))
        summary = class_report_summary(db, class_id)
        totals = db.execute(
            """
            SELECT
                COUNT(DISTINCT s.id) total_sessions,
                SUM(CASE WHEN a.status='present' THEN a.minutes ELSE 0 END) total_minutes
            FROM class_sessions s
            LEFT JOIN class_attendance a ON a.session_id = s.id
            WHERE s.class_id=?
            """,
            (class_id,),
        ).fetchone()
    return {"class": class_row, "summary": summary, "totals": dict(totals)}


@app.post("/api/classes/{class_id}/attendance-history")
def launch_attendance_history(
    class_id: int,
    data: AttendanceHistoryIn,
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    with connect() as db:
        class_row = get_class_or_404(db, class_id)
        get_student_or_404(db, data.user_id)
        start_iso = as_iso_date(data.start_date)
        end_iso = as_iso_date(data.end_date) if data.end_date else date.today().isoformat()
        rows = db.execute(
            """
            SELECT s.id session_id
            FROM class_sessions s
            WHERE s.class_id=? AND s.session_date BETWEEN ? AND ?
            ORDER BY s.session_date
            """,
            (class_id, start_iso, end_iso),
        ).fetchall()
        updated = 0
        for row in rows:
            minutes = class_row["default_minutes"] if data.status == "present" else 0
            db.execute(
                """
                INSERT INTO class_attendance
                (session_id,user_id,status,minutes,notes,registered_at)
                VALUES (?,?,?,?,?,?)
                ON CONFLICT(session_id,user_id) DO UPDATE SET
                status=excluded.status,
                minutes=excluded.minutes,
                notes=excluded.notes,
                registered_at=excluded.registered_at
                """,
                (row["session_id"], data.user_id, data.status, minutes, data.notes, now_iso()),
            )
            updated += 1
    return {"updated_sessions": updated, "status": data.status, "start_date": start_iso, "end_date": end_iso}


@app.get("/api/reports/classes")
def classes_report(
    program_id: int | None = None,
    period_label: str = "",
    academic_year: int | None = None,
    _: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    with connect() as db:
        where = ["c.active=1"]
        params: list[Any] = []
        if program_id:
            where.append("c.program_id=?")
            params.append(program_id)
        if period_label.strip():
            where.append("lower(c.period_label)=lower(?)")
            params.append(period_label.strip())
        if academic_year:
            where.append("c.academic_year=?")
            params.append(academic_year)
        classes = db.execute(
            f"""
            SELECT c.*, p.name program_name
            FROM class_groups c
            JOIN programs p ON p.id=c.program_id
            WHERE {' AND '.join(where)}
            ORDER BY p.name, c.weekday, c.start_time, c.name
            """,
            tuple(params),
        ).fetchall()
        result = []
        for class_row in classes:
            summary = class_report_summary(db, class_row["id"])
            result.append(
                {
                    "class": dict(class_row),
                    "summary": summary,
                    "student_count": len(summary),
                    "present_total": sum(item.get("presents") or 0 for item in summary),
                    "absence_total": sum(item.get("absences") or 0 for item in summary),
                    "minutes_total": sum(item.get("minutes") or 0 for item in summary),
                }
            )
    return {"classes": result}


@app.post("/api/sessions/{session_id}/attendance")
def save_session_attendance(session_id: int, data: AttendanceBulkIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    with connect() as db:
        session = db.execute("SELECT * FROM class_sessions WHERE id=?", (session_id,)).fetchone()
        if not session:
            raise HTTPException(404, "Aula não encontrada.")
        class_row = get_class_or_404(db, session["class_id"])
        for record in data.records:
            minutes = record.minutes if record.status == "present" else 0
            if record.status == "present" and minutes == 0:
                minutes = class_row["default_minutes"]
            db.execute(
                """
                INSERT INTO class_attendance
                (session_id,user_id,status,minutes,notes,registered_at)
                VALUES (?,?,?,?,?,?)
                ON CONFLICT(session_id,user_id) DO UPDATE SET
                status=excluded.status,
                minutes=excluded.minutes,
                notes=excluded.notes,
                registered_at=excluded.registered_at
                """,
                (session_id, record.user_id, record.status, minutes, record.notes, now_iso()),
            )
    return {"status": "ok", "saved": len(data.records)}


@app.get("/api/certificates/{student_id}", response_class=HTMLResponse)
def certificate(student_id: int, user: dict[str, Any] = Depends(current_user)) -> str:
    if user["role"] != "admin" and user["id"] != student_id:
        raise HTTPException(403, "Sem permissão.")
    with connect() as db:
        student = get_student_or_404(db, student_id)
        gym_minutes = db.execute("SELECT coalesce(sum(minutes),0) m FROM attendance WHERE user_id=?", (student_id,)).fetchone()["m"]
        class_minutes = db.execute("SELECT coalesce(sum(minutes),0) m FROM class_attendance WHERE user_id=? AND status='present'", (student_id,)).fetchone()["m"]
        minutes = gym_minutes + class_minutes
    hours = round(minutes / 60, 1)
    today = datetime.now().strftime("%d/%m/%Y")
    return f"""
    <!doctype html><html lang="pt-BR"><meta charset="utf-8">
    <title>Certificado CAFIS</title>
    <style>
      @page {{ size: A4 landscape; margin: 18mm; }}
      * {{ box-sizing: border-box; }}
      body {{ font-family: Arial, Helvetica, sans-serif; margin: 0; color: #17201a; background: #f5f7f6; }}
      .cert {{ border: 10px solid #0f7b4f; padding: 34px 46px; min-height: calc(100vh - 36mm); background: white; }}
      .header {{ display: flex; align-items: center; justify-content: space-between; gap: 24px; border-bottom: 2px solid #dce6e1; padding-bottom: 18px; }}
      .logo {{ width: 300px; max-width: 38%; height: auto; }}
      .org {{ text-align: right; font-size: 15px; line-height: 1.35; color: #405149; }}
      h1 {{ font-size: 42px; margin: 44px 0 28px; text-align: center; }}
      p {{ font-size: 22px; line-height: 1.65; }}
      .date {{ margin-top: 30px; text-align: right; }}
      .signatures {{ display: flex; justify-content: center; margin-top: 72px; }}
      .sig {{ border-top: 1px solid #333; width: 430px; text-align: center; padding-top: 10px; font-size: 16px; line-height: 1.35; }}
      .sig strong {{ display: block; font-size: 18px; }}
      @media print {{ body {{ background: white; }} .cert {{ min-height: auto; }} }}
    </style>
    <body><main class="cert">
      <header class="header">
        <img class="logo" src="/utfpr-logo.svg" alt="UTFPR">
        <div class="org">
          Universidade Tecnológica Federal do Paraná<br>
          Campus Ponta Grossa<br>
          CAFIS - Academia
        </div>
      </header>
      <h1>Certificado de Horas Complementares</h1>
      <p>Certificamos que <strong>{student['name']}</strong> participou das atividades da Academia CAFIS UTFPR Ponta Grossa, totalizando <strong>{hours} horas</strong> registradas por controle de entrada e saída.</p>
      <p class="date">Ponta Grossa, {today}.</p>
      <section class="signatures">
        <div class="sig">
          <strong>Prof. José Alves Faria Filho</strong>
          Chefe do CAFIS
        </div>
      </section>
    <script>window.print()</script></main></body></html>
    """


@app.get("/api/certificates/classes/{class_id}/students/{student_id}", response_class=HTMLResponse)
def class_certificate(class_id: int, student_id: int, user: dict[str, Any] = Depends(current_user)) -> str:
    if user["role"] != "admin" and user["id"] != student_id:
        raise HTTPException(403, "Sem permissão.")
    with connect() as db:
        class_row = get_class_or_404(db, class_id)
        student = get_student_or_404(db, student_id)
        enrolled = db.execute("SELECT id FROM enrollments WHERE class_id=? AND user_id=?", (class_id, student_id)).fetchone()
        if not enrolled:
            raise HTTPException(404, "Aluno não está matriculado nesta turma.")
        minutes = db.execute(
            """
            SELECT coalesce(sum(a.minutes),0) m
            FROM class_attendance a JOIN class_sessions s ON s.id=a.session_id
            WHERE s.class_id=? AND a.user_id=? AND a.status='present'
            """,
            (class_id, student_id),
        ).fetchone()["m"]
        presences = db.execute(
            """
            SELECT count(*) c FROM class_attendance a JOIN class_sessions s ON s.id=a.session_id
            WHERE s.class_id=? AND a.user_id=? AND a.status='present'
            """,
            (class_id, student_id),
        ).fetchone()["c"]
    hours = round(minutes / 60, 1)
    today = datetime.now().strftime("%d/%m/%Y")
    return f"""
    <!doctype html><html lang="pt-BR"><meta charset="utf-8">
    <title>Certificado CAFIS</title>
    <style>
      @page {{ size: A4 landscape; margin: 18mm; }}
      * {{ box-sizing: border-box; }}
      body {{ font-family: Arial, Helvetica, sans-serif; margin: 0; color: #17201a; background: #f5f7f6; }}
      .cert {{ border: 10px solid #0f7b4f; padding: 34px 46px; min-height: calc(100vh - 36mm); background: white; }}
      .header {{ display: flex; align-items: center; justify-content: space-between; gap: 24px; border-bottom: 2px solid #dce6e1; padding-bottom: 18px; }}
      .logo {{ width: 300px; max-width: 38%; height: auto; }}
      .org {{ text-align: right; font-size: 15px; line-height: 1.35; color: #405149; }}
      h1 {{ font-size: 42px; margin: 40px 0 26px; text-align: center; }}
      p {{ font-size: 22px; line-height: 1.65; }}
      .date {{ margin-top: 30px; text-align: right; }}
      .signatures {{ display: flex; justify-content: center; margin-top: 72px; }}
      .sig {{ border-top: 1px solid #333; width: 430px; text-align: center; padding-top: 10px; font-size: 16px; line-height: 1.35; }}
      .sig strong {{ display: block; font-size: 18px; }}
    </style>
    <body><main class="cert">
      <header class="header">
        <img class="logo" src="/utfpr-logo.svg" alt="UTFPR">
        <div class="org">Universidade Tecnológica Federal do Paraná<br>Campus Ponta Grossa<br>CAFIS</div>
      </header>
      <h1>Certificado de Participação</h1>
      <p>Certificamos que <strong>{student['name']}</strong> participou da turma <strong>{class_row['name']}</strong>, vinculada ao projeto <strong>{class_row['program_name']}</strong>, totalizando <strong>{hours} horas</strong> e <strong>{presences} presença(s)</strong> registradas no sistema CAFIS.</p>
      <p class="date">Ponta Grossa, {today}.</p>
      <section class="signatures"><div class="sig"><strong>Prof. José Alves Faria Filho</strong>Chefe do CAFIS</div></section>
    <script>window.print()</script></main></body></html>
    """


app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")
